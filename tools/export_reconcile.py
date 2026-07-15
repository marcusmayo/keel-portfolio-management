#!/usr/bin/env python3
"""Export reconcile.json to a human-readable .xlsx worklist, one sheet per bucket.
Reads state/normalized/reconcile.json, writes exports/reconcile-<date>.xlsx.
Prints the output path (the webchat serves it for download)."""

import json, sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = Path("state/normalized/reconcile.json")
BACKLOG = Path("state/normalized/backlog.json")
OUTDIR = Path("exports")

# column order shown in every sheet (keys map into each row dict; missing -> blank)
COLS = [
    ("src_name",    "Backlog Item"),
    ("src_ref",     "Source Ref"),
    ("type",        "Type"),
    ("src_status",  "Backlog Status"),
    ("keel_key",    "Keel Key"),
    ("keel_name",   "Keel Item"),
    ("keel_status", "Keel Status"),
    ("wsjf",        "WSJF"),
    ("rice",        "RICE"),
    ("score",       "Match"),
    ("verdict",          "Verdict"),
    ("semantic_verdict", "Semantic"),
    ("semantic_reason",  "Semantic Reason"),
    ("reason",           "Reason"),
    ("action",           "Proposed Action"),
]

# bucket order (sheet tabs); skip empty buckets so the file isn't cluttered
BUCKET_ORDER = ["completed", "conflict", "ambiguous", "changed", "duplicate", "gap", "done_gap"]

HEADER_FILL = PatternFill("solid", fgColor="1F3B1B")
HEADER_FONT = Font(bold=True, color="FFFFFF")

def write_sheet(wb, name, rows):
    ws = wb.create_sheet(title=name[:31])
    # header
    for c, (_, label) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    # rows
    for r, row in enumerate(rows, 2):
        for c, (key, _) in enumerate(COLS, 1):
            v = row.get(key, "")
            ws.cell(row=r, column=c, value=("" if v is None else v))
    # column widths (cap so Reason/Item don't run away)
    widths = {"src_name": 42, "keel_name": 38, "reason": 40, "action": 26, "semantic_reason": 48, "semantic_verdict": 11,
              "src_status": 16, "keel_status": 14}
    for c, (key, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(key, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
    return ws

import subprocess, glob as _glob
from _require import require

def _git_short():
    try:
        return subprocess.check_output(["git", "rev-parse", "--short", "HEAD"],
                                       text=True).strip() or "nogit"
    except Exception:
        return "nogit"

def load_support():
    """Read support/*.yaml (note-derived feature/bug PROPOSALS) for the Proposals tab.
    Visibility only - these are not portfolio items until promoted to Jira."""
    import yaml
    out = []
    for f in sorted(_glob.glob("support/*.yaml")):
        try:
            w = (yaml.safe_load(open(f, encoding="utf-8")) or {}).get("workitem") or {}
        except Exception:
            continue
        if not w:
            continue
        out.append({
            "key": w.get("key", ""), "type": w.get("type", ""),
            "name": w.get("name", ""), "status": w.get("status", ""),
            "origin": (w.get("source") or {}).get("origin", ""),
            "next_action": w.get("next_action", ""),
        })
    return out

def write_proposals(wb, props):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(title="Proposals")
    cols = [("Key", 12), ("Type", 8), ("Name", 50), ("Status", 12),
            ("Origin", 10), ("Next Action", 46)]
    for c, (label, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3B1B")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    keys = ["key", "type", "name", "status", "origin", "next_action"]
    for r, row in enumerate(props, 2):
        for c, k in enumerate(keys, 1):
            ws.cell(row=r, column=c, value=row.get(k, ""))
    ws.freeze_panes = "A2"
    return ws

def write_merge(wb, proposals):
    """Keel-origin -> Jira merge proposals (from merge_pass). Review surface: confirm via
    /merge-accept in webchat. Flags story/bug targets claimed by more than one keel item."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(title="Merge")
    cols = [("Keel Key", 10), ("Keel Type", 9), ("Keel Item", 40), ("->", 4),
            ("Jira Ref", 10), ("Jira Item", 40), ("SAME Reason", 52), ("Flag", 22)]
    for c, (label, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B2E3B")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    # detect many-to-one on a jira_ref (flagged only for story/bug targets by caller data)
    from collections import Counter
    ref_counts = Counter(p.get("jira_ref","") for p in proposals)
    warn = PatternFill("solid", fgColor="5A2D2D")
    for r, pr in enumerate(proposals, 2):
        ref = pr.get("jira_ref","")
        many = ref_counts[ref] > 1
        vals = [pr.get("keel_key",""), pr.get("keel_type",""), pr.get("keel_name",""), "->",
                ref, pr.get("jira_name",""), pr.get("reason",""),
                ("CONFLICT: ref claimed x%d" % ref_counts[ref]) if many else ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if many:
                cell.fill = warn
    ws.freeze_panes = "A2"
    return ws

def write_unknown_types(wb, rows):
    """Backlog rows the source sheet left with empty/invalid Type - operator
    review surface. Rule via: python3 tools/set_backlog_type.py TASK#=type
    then re-run /normalize. Sheet self-clears as rulings land."""
    ws = wb.create_sheet(title="Unknown Type")
    cols = ["Task #", "Backlog Item", "Backlog Status", "Priority", "Source Ref",
            "Notes", "Ruling (epic/feature/story)"]
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    r = 1
    for r, row in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=row.get("task_id", ""))
        ws.cell(row=r, column=2, value=row.get("name", ""))
        ws.cell(row=r, column=3, value=row.get("raw_status", ""))
        ws.cell(row=r, column=4, value=row.get("priority", ""))
        ws.cell(row=r, column=5, value=(row.get("source") or {}).get("ref", ""))
        ws.cell(row=r, column=6, value=row.get("description", ""))
    ws.cell(row=r + 2, column=1,
            value="Apply rulings: python3 tools/set_backlog_type.py TASK#=type TASK#=type ...  then re-run /normalize")
    widths = {1: 8, 2: 44, 3: 14, 4: 12, 5: 10, 6: 60, 7: 26}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(r,1)+1}"
    return ws


def main():
    data = json.loads(require(SRC).read_text(encoding="utf-8"))
    try:
        _bl = json.loads(require(BACKLOG).read_text(encoding="utf-8"))
        unknown_rows = [r for r in _bl.get("rows", []) if r.get("type") == "unknown"]
    except FileNotFoundError:
        unknown_rows = []
    buckets = data["buckets"]
    summary = data["summary"]

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    # summary tab first
    ws = wb.create_sheet(title="Summary")
    ws.cell(row=1, column=1, value="Reconcile Summary").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {data.get('generated','')}")
    ws.cell(row=3, column=1, value=f"Backlog rows: {summary.get('backlog_rows','')}   "
                                   f"Keel items: {summary.get('keel_items','')}")
    _stamp = datetime.now()
    ws.cell(row=4, column=1, value=f"Exported: {_stamp.strftime('%Y-%m-%d %H:%M')}   "
                                   f"matcher version: {_git_short()}")
    r = 5
    ws.cell(row=r, column=1, value="Bucket").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Count").font = Font(bold=True)
    for b in BUCKET_ORDER:
        r += 1
        ws.cell(row=r, column=1, value=b)
        ws.cell(row=r, column=2, value=len(buckets.get(b, [])))
    ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 10
    if unknown_rows:
        r += 2
        ws.cell(row=r, column=1, value="Unknown type (backlog)").font = Font(bold=True)
        ws.cell(row=r, column=2, value=len(unknown_rows))

    # one sheet per non-empty bucket
    made = []
    for b in BUCKET_ORDER:
        rows = buckets.get(b, [])
        if rows:
            write_sheet(wb, b.capitalize(), rows)
            made.append(f"{b}={len(rows)}")

    # note-derived proposals (support/) as a labeled review tab
    props = load_support()
    write_proposals(wb, props)
    made.append(f"proposals={len(props)}")

    if unknown_rows:
        write_unknown_types(wb, unknown_rows)
        made.append(f"unknown_type={len(unknown_rows)}")

    # keel-origin -> jira merge proposals (from merge_pass, if present)
    merges = data.get("buckets", {}).get("merge_candidate", [])
    if merges:
        write_merge(wb, merges)
        made.append(f"merge={len(merges)}")

    OUTDIR.mkdir(parents=True, exist_ok=True)
    _ts = datetime.now()
    out = OUTDIR / f"reconcile-{_ts.strftime('%Y-%m-%d')}_{_ts.strftime('%H%M')}_{_git_short()}.xlsx"
    wb.save(out)
    print(f"wrote {out}  ({'  '.join(made)})")
    print(str(out))  # last line = path, for the server route to locate

if __name__ == "__main__":
    main()
