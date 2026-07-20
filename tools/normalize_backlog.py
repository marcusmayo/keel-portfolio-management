#!/usr/bin/env python3
"""Deterministic backlog normalizer. Reads the newest raw backlog xlsx
(knowledge/import/raw/*Backlog*.xlsx; argv[1] overrides), maps each row to a
canonical work-item dict, writes JSON for reconcile. openpyxl for xlsx read;
no LLM. Prints a counts summary for operator review vs prior run."""

import csv, re, json, sys, io
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "knowledge" / "import" / "raw"
OUT = ROOT / "state" / "normalized" / "backlog.json"
OVR_PATH = ROOT / "state" / "backlog-type-overrides.json"
try:
    OVERRIDES = json.loads(OVR_PATH.read_text(encoding="utf-8")).get("overrides", {})
except FileNotFoundError:
    OVERRIDES = {}

# --- confirmed status map (NYS=not-started, IP=in-progress) ---
STATUS_MAP = {
    "DONE": "done",
    "NYS": "not-started",
    "IP": "in-progress",
    "DUPLICATE": "dedup-flag",
    "IN ANALYSIS": "analysis",
    "NEEDS ANALYSIS": "analysis",
    "BLOCKED": "blocked",
    "": "unscored",
}
TYPE_VOCAB = {"epic", "feature", "story"}
def _load_source_prefix():
    """SOURCE_KEY_PREFIX from $KEEL_CONFIG if set, else keel.config.json at the repo root; None if absent/unset.
    Free-text key scanning needs a declared prefix (generic [A-Z]+-\\d+
    would false-positive on ISO-9001-style tokens)."""
    import json as _json
    from pathlib import Path as _P
    import os as _os
    _envcfg = _os.environ.get("KEEL_CONFIG")
    cfg = _P(_envcfg) if _envcfg else _P(__file__).resolve().parent.parent / "keel.config.json"
    try:
        return (_json.loads(cfg.read_text(encoding="utf-8")).get("SOURCE_KEY_PREFIX") or None)
    except Exception:
        return None
SOURCE_KEY_PREFIX = _load_source_prefix()
NGE_RE = re.compile(re.escape(SOURCE_KEY_PREFIX) + r"-\d+") if SOURCE_KEY_PREFIX else None

def load_csv_block(path):
    """Strip frontmatter, find the ### Sheet: line, return CSV text after it."""
    text = path.read_text(encoding="utf-8")
    lines = text.split("\n")
    # skip frontmatter (--- ... ---)
    i = 0
    if lines and lines[0].strip() == "---":
        i = 1
        while i < len(lines) and lines[i].strip() != "---":
            i += 1
        i += 1  # past closing ---
    # find ### Sheet: header
    while i < len(lines) and not lines[i].startswith("### Sheet:"):
        i += 1
    i += 1  # the CSV starts on the next line
    return "\n".join(lines[i:])

def latest_backlog_xlsx():
    """Newest date-stamped backlog xlsx in knowledge/import/raw/."""
    files = sorted(RAW_DIR.glob("*Backlog*.xlsx"))
    if not files:
        sys.exit(f"ERROR: no *Backlog*.xlsx in {RAW_DIR}")
    return files[-1]

def load_xlsx_rows(path):
    """Read the raw backlog xlsx; return dicts keyed by header row.
    Sheet chosen by 'Task #' in row 1 (header-detected, ignores junk sheets).
    Values stringified; None -> ''; integral floats -> int string."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        first = next(wb[name].iter_rows(max_row=1, values_only=True), ())
        if any(str(v).strip() == "Task #" for v in first if v is not None):
            ws = wb[name]
            break
    if ws is None:
        sys.exit(f"ERROR: no sheet with 'Task #' header in {path}")
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    out = []
    for row in rows_iter:
        d = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            if v is None:
                d[h] = ""
            elif isinstance(v, float) and v.is_integer():
                d[h] = str(int(v))
            else:
                d[h] = str(v)
        out.append(d)
    return out

def clean(v):
    """Trim, strip nbsp placeholders -> empty."""
    if v is None:
        return ""
    v = v.replace("\u00a0", " ").strip()
    return "" if v in ("", "-") else v

def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_backlog_xlsx()
    reader = load_xlsx_rows(src)

    rows = []
    flags = {"unknown_type": [], "unknown_status": [], "dedup": [], "nge_refs": [], "override_applied": [], "override_stale": [], "override_redundant": []}
    type_counts = {"epic": 0, "feature": 0, "story": 0, "unknown": 0}

    for n, raw in enumerate(reader, start=1):
        name   = clean(raw.get("Feature"))
        rawtype = clean(raw.get("Type"))
        rawstat = clean(raw.get("Status"))
        weeks  = clean(raw.get("Weeks"))
        prio   = clean(raw.get("Priority"))
        notes  = clean(raw.get("Notes"))
        taskid = clean(raw.get("Task #"))

        if not name and not rawtype and not rawstat and not notes:
            continue  # genuinely empty trailing row

        # type: never guess
        t = rawtype.lower()
        ovr = OVERRIDES.get(taskid)
        if t in TYPE_VOCAB:
            wtype = t
            if ovr:
                mark = " (CONFLICTS with source)" if ovr.get("type") != t else " (matches source; removable)"
                flags["override_redundant"].append(f"row {n} '{taskid}': source type now {t!r}; ruling {ovr.get('type')!r}{mark}")
        elif ovr and ovr.get("name", "").strip().lower() == name.strip().lower():
            wtype = ovr["type"]
            flags["override_applied"].append(f"row {n} '{taskid}': -> {ovr['type']}")
        elif ovr:
            wtype = "unknown"
            flags["override_stale"].append(f"row {n} '{taskid}': name changed since ruling (was {ovr.get('name','')!r}) - skipped")
            flags["unknown_type"].append(f"row {n} '{taskid}': stale override, name={name!r}")
        else:
            wtype = "unknown"
            if rawtype:  # non-empty but unrecognized (e.g. "Himanshu")
                flags["unknown_type"].append(f"row {n} '{taskid}': type={rawtype!r} name={name!r}")
            else:
                flags["unknown_type"].append(f"row {n} '{taskid}': type=EMPTY name={name!r}")
        type_counts[wtype] += 1

        # status
        skey = rawstat.upper()
        if skey in STATUS_MAP:
            wstat = STATUS_MAP[skey]
        else:
            wstat = "unknown"
            flags["unknown_status"].append(f"row {n} '{taskid}': status={rawstat!r}")
        if wstat == "dedup-flag":
            flags["dedup"].append(f"row {n} '{taskid}': {name!r}")

        # embedded source key -> proposed ref (skipped when no prefix configured)
        refs = NGE_RE.findall(notes) if NGE_RE else []
        ref = refs[0] if refs else ""
        if refs:
            flags["nge_refs"].append(f"row {n} '{taskid}': {refs} (name={name!r})")

        rows.append({
            "type": wtype,
            "name": name,
            "status": wstat,
            "raw_status": rawstat,
            "effort_weeks": weeks,
            "priority": prio,
            "description": ("[draft - review] " + notes) if notes else "",
            "task_id": taskid,
            "parent": "",
            "source": {"origin": "backlog-xlsx", "ref": ref},
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "generated": datetime.now().astimezone().isoformat(),
        "source_file": str(src),
        "row_count": len(rows),
        "type_counts": type_counts,
        "rows": rows,
    }, indent=2), encoding="utf-8")

    # --- summary to chat ---
    print(f"=== normalize_backlog: {len(rows)} rows ===")
    if OVERRIDES:
        print(f"type overrides: applied={len(flags['override_applied'])}  stale={len(flags['override_stale'])}  redundant={len(flags['override_redundant'])}  (store: {len(OVERRIDES)})")
        for line in flags["override_applied"]: print("    applied:", line)
        for line in flags["override_stale"]: print("    STALE:", line)
        for line in flags["override_redundant"]: print("    redundant:", line)
    print(f"type counts: epic={type_counts['epic']}  feature={type_counts['feature']}  "
          f"story={type_counts['story']}  unknown={type_counts['unknown']}")
    print(f"  (known-good target: epic=38  feature=1  story=33  unknown=33  total=92? -> {len(rows)})")
    print(f"dedup-flagged (source-declared Duplicate): {len(flags['dedup'])}")
    for d in flags["dedup"]: print("   ", d)
    print(f"embedded source refs (proposed, never auto-linked): {len(flags['nge_refs'])}")
    if NGE_RE is None:
        print("(embedded source-key scan skipped: no SOURCE_KEY_PREFIX in keel.config.json)")
    for d in flags["nge_refs"][:8]: print("   ", d)
    if len(flags["nge_refs"]) > 8: print(f"    ... +{len(flags['nge_refs'])-8} more")
    print(f"unknown status (needs operator decision): {len(flags['unknown_status'])}")
    for d in flags["unknown_status"][:8]: print("   ", d)
    print(f"unknown type (empty or non-vocab, NOT guessed): {len(flags['unknown_type'])}")
    for d in flags["unknown_type"][:8]: print("   ", d)
    if len(flags["unknown_type"]) > 8: print(f"    ... +{len(flags['unknown_type'])-8} more")
    print(f"wrote {OUT}")

if __name__ == "__main__":
    main()
