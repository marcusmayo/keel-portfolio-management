#!/usr/bin/env python3
"""Executable oracle for the Northwind E2E (RCA-1 corrective).
Run from the tree root after:
  gen_corpus.py .  ->  normalize_jira  ->  normalize_backlog  ->  apply.py --commit  ->  reconcile.py backlog  ->  export_multisource.py
Read-only. Exit 0 = pass, 2 = fail with named diffs."""
import glob, json, sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _recon import load

def need(p, producer):
    if not Path(p).exists():
        print(f"ERROR: missing {p} -- produce with: {producer}")
        sys.exit(2)

FAILS = []
def check(name, got, want):
    ok = got == want
    print(f"  {'PASS' if ok else 'FAIL'}  {name}: {got!r}" + ("" if ok else f"  (want {want!r})"))
    if not ok:
        FAILS.append(name)

need("expectations.json", "python3 examples/northwind/gen_corpus.py .")
need("state/normalized/reconcile.json", "python3 tools/reconcile.py backlog")
exp = json.loads(Path("expectations.json").read_text(encoding="utf-8"))
E = exp["backlog_reconcile"]
_d = load()
_lane = (_d.get("summary") or {}).get("source")
if _lane != "backlog":
    print(f"ERROR: reconcile.json is the {_lane!r} lane (export reruns jira last) -- re-run: python3 tools/reconcile.py backlog")
    sys.exit(2)
b = _d["buckets"]

for k, want in E["buckets"].items():
    check(f"bucket {k}", len(b.get(k, [])), want)

ch = b["changed"]
check("match_mode ref", sum(1 for x in ch if x.get("match_mode") == "ref"), E["match_modes"]["ref"])
check("match_mode title", sum(1 for x in ch if x.get("match_mode") == "title"), E["match_modes"]["title"])
check("gap src_names", sorted(x.get("src_name", "") for x in b["gap"]), sorted(E["gap_src_names"]))

P = E["plants"]
by_ref = {x.get("src_ref"): x for x in ch if x.get("src_ref")}
sd = by_ref.get(P["status_drift"]["src_ref"], {})
check("status-drift src_status", sd.get("src_status"), P["status_drift"]["src_status"])
check("status-drift keel_status", sd.get("keel_status"), P["status_drift"]["keel_status"])
td = by_ref.get(P["title_drift"]["src_ref"], {})
check("title-drift src_name", td.get("src_name"), P["title_drift"]["src_name"])
check("title-drift keel_name", td.get("keel_name"), P["title_drift"]["keel_name"])
check("embedded-key match_mode", by_ref.get(P["embedded_key"]["src_ref"], {}).get("match_mode"),
      P["embedded_key"]["match_mode"])
nd = [x for x in ch if x.get("match_mode") == "title"]
nd0 = nd[0] if nd else {}
check("near-dup src_name", nd0.get("src_name"), P["near_dup"]["src_name"])
paired = by_ref.get(P["near_dup"]["shares_item_with_ref"], {})
check("near-dup multi-claim", bool(nd0.get("keel_key")) and nd0.get("keel_key") == paired.get("keel_key"), True)

amb = b["ambiguous"]
ap0 = next((x for x in amb if x.get("src_name") == P["ambiguous_paraphrase"]["src_name"]), {})
check("ambiguous-paraphrase src_name", ap0.get("src_name"), P["ambiguous_paraphrase"]["src_name"])
check("ambiguous-paraphrase keel_key", ap0.get("keel_key"), P["ambiguous_paraphrase"]["keel_key"])

import yaml
yams = sorted(glob.glob("state/*.yaml"))
check("state item count", len(yams), exp["state_items"]["count"])
state_keys = set()
for f in yams:
    try:
        state_keys.add(((yaml.safe_load(open(f, encoding="utf-8")) or {}).get("workitem") or {}).get("key"))
    except Exception as e:
        FAILS.append(f"yaml parse {f}")
        print(f"  FAIL  yaml parse {f}: {e}")
check("changed keel_keys all exist in state/",
      sorted(k for k in {x.get("keel_key") for x in ch} if k not in state_keys), [])


# --- export layer ---
import glob as _g
xs = sorted(_g.glob("exports/multisource-*.xlsx"))
if not xs:
    print("ERROR: missing exports/multisource-*.xlsx -- produce with: python3 tools/export_multisource.py")
    sys.exit(2)
import openpyxl
EX = exp["export"]
wb = openpyxl.load_workbook(xs[-1], read_only=True)
check("export sheets", wb.sheetnames, EX["sheets"])
def sheet_rows(name):
    rows = list(wb[name].iter_rows(values_only=True))
    hdr = list(rows[0]) if rows else []
    return [dict(zip(hdr, r)) for r in rows[1:] if any(c not in (None, "") for c in r)]
cs = sheet_rows("Cross-Source")
check("cross-source rows", len(cs), EX["cross_source_rows"])
dist = {}
for d in cs: dist[str(d.get("Sources"))] = dist.get(str(d.get("Sources")), 0) + 1
check("sources dist", dist, EX["sources_dist"])
check("disagree YES jira keys",
      sorted(str(d.get("Jira Key")) for d in cs if str(d.get("Disagree?") or "") == "YES"),
      EX["disagree_yes_jira_keys"])
for jk in EX["no_disagree_jira_keys"]:
    row = [d for d in cs if str(d.get("Jira Key")) == jk]
    check(f"{jk} present+unflagged", bool(row) and not (row[0].get("Disagree?") or ""), True)
check("source-only rows", len(sheet_rows("Source-Only")), EX["source_only_rows"])
check("keel-origin rows", len(sheet_rows("Keel-Origin")), EX["keel_origin_rows"])
check("unconfirmed rows", len(sheet_rows("Unconfirmed")), EX["unconfirmed_rows"])
wb.close()
print()
if FAILS:
    print(f"VERIFY: FAIL ({len(FAILS)}): {', '.join(FAILS)}")
    sys.exit(2)
print("VERIFY: ALL PASS")
