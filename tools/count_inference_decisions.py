#!/usr/bin/env python3
"""Print the number of confirm/reject decisions in an Unconfirmed sheet.
Used by the webchat to route inference uploads. Prints 0 on any problem
(diagnostic traceback to stderr). Extension not required on the path."""
import sys
try:
    from openpyxl import load_workbook
    wb = load_workbook(open(sys.argv[1], "rb"), read_only=True)
    n = 0
    if "Unconfirmed" in wb.sheetnames:
        ws = wb["Unconfirmed"]
        h = [str(c.value).strip() if c.value else "" for c in ws[1]]
        di = next((i for i, x in enumerate(h) if x.startswith("Decision")), None)
        ki = next((i for i, x in enumerate(h) if x == "Unconfirmed Key"), None)
        if di is not None and ki is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[ki] is None:
                    continue
                v = str(r[di]).strip().lower() if r[di] not in (None, "") else ""
                if v in ("confirm", "reject"):
                    n += 1
    print(n)
except Exception:
    import traceback
    traceback.print_exc(file=sys.stderr)
    print(0)
