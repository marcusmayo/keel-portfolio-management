#!/usr/bin/env python3
"""Round-trip: apply operator edits from a portfolio/multi-source export back to
the YAMLs. Approval signal is removal of the "[draft — review]" prefix in the
sheet (operator's rule): a Description whose prefix is gone AND value changed is
written; an AC line whose prefix is gone is written approved, line by line,
while still-prefixed lines stay draft.

Keel-owned fields ONLY: description, acceptance_criteria. Never touches ref,
Jira-origin status, type, parent, hierarchy. Match by the item's key: field.
DRY-RUN by default; --commit writes. .pre-edit backup per changed file.

Accepts either export layout (detected by header columns):
  - portfolio export  (keel-portfolio-*): "Keel Key", "Description", "Acceptance Criteria..."
  - multi-source KO    (keel-multisource-*): "Keel Key", "Description (draft)", "Acceptance Criteria (draft)"
"""
import re, sys, glob
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
STATE = ROOT / "state"
SUPPORT = ROOT / "support"
DRAFT = "[draft — review]"
# Tolerate any dash variant (em/en/hyphen) + flexible spacing in the draft marker,
# so byte-level dash differences between files never break detection.
_DRAFT_RE = re.compile(r"^\s*\[\s*draft\s*[—–-]\s*review\s*\]\s*")

def find_sheet_and_cols(wb):
    """Locate the editable sheet and its key/desc/ac columns. Accepts either
    export vocabulary: portfolio export (Key/Description/AcceptanceCriteria) or
    multi-source (Keel Key/Description (draft)/Acceptance Criteria (draft)).
    Requires BOTH a key column AND a description column, so the portfolio
    workbook's Roadmap sheet (key but no description) is skipped in favour of
    the Portfolio sheet."""
    want_key = {"keel key", "key"}
    want_desc = {"description", "description (draft)"}
    want_ac = {"acceptancecriteria", "acceptance criteria",
               "acceptance criteria (draft)", "acceptance criteria (epic/feature/story)"}
    for ws in wb.worksheets:
        hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        def idx(cands):
            for i, h in enumerate(hdr):
                if h in cands:
                    return i
            return None
        ki, di, ai = idx(want_key), idx(want_desc), idx(want_ac)
        if ki is not None and di is not None:   # need key AND description to be editable
            return ws, ki, di, ai
    return None, None, None, None

def key_to_file():
    """Map each item's key: field -> its YAML path (state/ + support/)."""
    m = {}
    for d in (STATE, SUPPORT):
        for f in glob.glob(str(d / "*.yaml")):
            if Path(f).stem.startswith("_"):
                continue
            txt = Path(f).read_text(encoding="utf-8")
            km = re.search(r"^\s*key:\s*([A-Z]+-\d+)", txt, re.M)
            if km:
                m[km.group(1)] = f
    return m

def yaml_get_desc(txt):
    m = re.search(r'^  description:\s*"(.*)"\s*$', txt, re.M)
    return m.group(1) if m else None

def yaml_get_ac_lines(txt):
    """Return the list of AC item strings (without the '- ' / quotes)."""
    m = re.search(r"^  acceptance_criteria:\s*(\[\]|\n((?:    -.*\n?)+))", txt, re.M)
    if not m:
        return None
    if m.group(1).strip() == "[]":
        return []
    items = []
    for line in m.group(2).splitlines():
        lm = re.match(r'\s*-\s*"(.*)"\s*$', line)
        if lm:
            items.append(lm.group(1))
    return items

def esc(s):
    return s.replace("\\", "\\\\").replace('"', '\\"')

def serialize_ac(items):
    if not items:
        return "  acceptance_criteria: []"
    out = ["  acceptance_criteria:"]
    for it in items:
        out.append(f'    - "{esc(it)}"')
    return "\n".join(out)

def strip_prefix(s):
    s = (s or "").strip()
    m = _DRAFT_RE.match(s)
    if m:
        return s[m.end():].strip(), True   # (clean, was_prefixed)
    return s, False

def main():
    if len(sys.argv) < 2 or sys.argv[1].startswith("--"):
        # allow: tool <xlsx> [--commit]  OR  tool --commit <xlsx>
        args = [a for a in sys.argv[1:] if not a.startswith("--")]
    else:
        args = [sys.argv[1]]
    commit = "--commit" in sys.argv
    if not args:
        # default to newest keel-portfolio/multisource in exports/
        cands = sorted(glob.glob(str(ROOT / "exports" / "keel-portfolio-*.xlsx")) +
                       glob.glob(str(ROOT / "exports" / "multisource-*.xlsx")) +
                       glob.glob(str(ROOT / "exports" / "keel-multisource-*.xlsx")))
        if not cands:
            sys.exit("ERROR: no xlsx given and none found in exports/")
        xlsx = cands[-1]
    else:
        xlsx = args[0]

    wb = load_workbook(xlsx, read_only=True)
    ws, ki, di, ai = find_sheet_and_cols(wb)
    if ws is None or ki is None:
        sys.exit("ABORT: no sheet with a 'Keel Key' column - not a portfolio/multi-source export")
    print(f"source: {xlsx}")
    print(f"sheet: {ws.title} | key col {ki} | desc col {di} | ac col {ai}")

    k2f = key_to_file()
    changed_desc = changed_ac_lines = unchanged = still_draft = no_match = jira_skipped = 0
    plan = []  # (file, key, new_desc_or_None, new_ac_items_or_None, notes)

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[ki]).strip() if ki is not None and row[ki] else ""
        if not key:
            continue
        f = k2f.get(key)
        if not f:
            no_match += 1
            continue
        txt = Path(f).read_text(encoding="utf-8")
        new_desc = None
        new_ac = None
        notes = []

        # HARD GUARD: Jira-origin items (have a source ref) are Jira-owned - their
        # description/AC live in Jira and are never authored via the Keel round-trip.
        # Skip regardless of what the sheet shows (even a de-prefixed row).
        if re.search(r"^\s*source:.*ref:\s*[A-Z]+-\d+", txt, re.M):
            jira_skipped += 1
            continue

        # --- Description: approve ONLY on prefix TRANSITION (YAML had prefix, sheet removed it) ---
        if di is not None and row[di] is not None:
            sheet_desc = str(row[di])
            clean, sheet_had_pref = strip_prefix(sheet_desc)
            cur = yaml_get_desc(txt)
            _, yaml_had_pref = strip_prefix(cur or "")
            if yaml_had_pref and not sheet_had_pref and clean:
                # the draft prefix existed and the operator removed it -> approve
                new_desc = clean
                changed_desc += 1
                notes.append("desc approved")
            elif yaml_had_pref and sheet_had_pref:
                still_draft += 1
            # yaml never had a prefix (already-authored) -> not an approval, ignore

        # --- AC: line-level, approve a line ONLY on prefix transition (YAML line had prefix, sheet removed it) ---
        if ai is not None and row[ai] is not None:
            sheet_lines = [l for l in str(row[ai]).split("\n") if l.strip()]
            cur_ac = yaml_get_ac_lines(txt) or []
            # map YAML lines by their de-prefixed clean text -> whether YAML had prefix
            yaml_pref = {}
            for yl in cur_ac:
                c, hp = strip_prefix(yl)
                yaml_pref[c] = hp
            rebuilt = []
            approved_here = 0
            for line in sheet_lines:
                clean, sheet_had_pref = strip_prefix(line)
                yaml_had_pref = yaml_pref.get(clean, None)
                if yaml_had_pref and not sheet_had_pref:
                    rebuilt.append(clean)          # prefix existed and was removed -> approve
                    approved_here += 1
                elif sheet_had_pref:
                    rebuilt.append(f"{DRAFT} {clean}")   # still draft
                else:
                    # line never had a prefix in YAML (already authored) -> keep as-is
                    rebuilt.append(clean)
            # only stage AC if a real approval transition happened
            if approved_here > 0 and rebuilt != cur_ac:
                new_ac = rebuilt
                changed_ac_lines += approved_here
                notes.append(f"ac: {approved_here} line(s) approved")

        if new_desc is not None or new_ac is not None:
            plan.append((f, key, new_desc, new_ac, "; ".join(notes)))
        else:
            unchanged += 1

    print(f"\n{'COMMIT' if commit else 'DRY-RUN'} summary:")
    print(f"  rows with description approved: {changed_desc}")
    print(f"  AC lines approved:              {changed_ac_lines}")
    print(f"  rows still fully draft:         {still_draft}")
    print(f"  rows unchanged:                 {unchanged}")
    print(f"  sheet keys not in portfolio:    {no_match}")
    print(f"  Jira-origin (skipped, Jira-owned): {jira_skipped}")
    print(f"  files to write:                 {len(plan)}")
    print("\nsample changes (first 10):")
    for f, key, nd, nac, notes in plan[:10]:
        print(f"  {key}: {notes}")

    if not commit:
        print("\nDRY-RUN: nothing written. Re-run with --commit to apply.")
        return

    written = 0
    for f, key, nd, nac, notes in plan:
        p = Path(f)
        txt = p.read_text(encoding="utf-8")
        orig = txt
        if nd is not None:
            txt = re.sub(r'^  description:\s*".*"\s*$',
                         f'  description: "{esc(nd)}"', txt, count=1, flags=re.M)
        if nac is not None:
            block = serialize_ac(nac)
            txt = re.sub(r"^  acceptance_criteria:\s*(\[\]|\n(?:    -.*\n?)+)",
                         block + "\n", txt, count=1, flags=re.M)
        if txt != orig:
            p.with_suffix(p.suffix + ".pre-edit").write_text(orig, encoding="utf-8")
            p.write_text(txt, encoding="utf-8")
            written += 1
    print(f"\nCOMMITTED: wrote {written} files (.pre-edit backups).")

if __name__ == "__main__":
    main()
