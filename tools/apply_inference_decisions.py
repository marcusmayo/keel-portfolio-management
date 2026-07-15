#!/usr/bin/env python3
"""apply_inference_decisions.py -- round-trip apply for the Unconfirmed sheet.

Reads a multisource export's Unconfirmed sheet, and for each row where the
Decision column is 'confirm' or 'reject', applies that decision by reusing the
proven confirm_item.py logic. DRY-RUN default; --commit writes with .pre-confirm
backups + hash-chained audit entries. Blank Decision = skipped (pending).

Usage:
  python3 tools/apply_inference_decisions.py path/to/multisource.xlsx
  python3 tools/apply_inference_decisions.py path/to/multisource.xlsx --commit
"""
import argparse, glob, json, re, shutil, subprocess, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import confirm_item as ci  # reuse apply_confirm / apply_reject / audit_record / files / key_of / inferred_flag

def load_decisions(xlsx):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True)
    if "Unconfirmed" not in wb.sheetnames:
        sys.exit("ABORT: workbook has no 'Unconfirmed' sheet")
    ws = wb["Unconfirmed"]
    hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
    hi = {h: i for i, h in enumerate(hdr)}
    if "Unconfirmed Key" not in hi:
        sys.exit("ABORT: no 'Unconfirmed Key' column")
    dcol = next((h for h in hdr if h.startswith("Decision")), None)
    if not dcol:
        sys.exit("ABORT: no Decision column")
    ki, di = hi["Unconfirmed Key"], hi[dcol]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ki] is None:
            continue
        dec = (str(row[di]).strip().lower() if row[di] not in (None, "") else "")
        if dec in ("confirm", "reject"):
            out.append((str(row[ki]).strip(), dec))
        elif dec:
            out.append((str(row[ki]).strip(), f"__invalid:{dec}"))
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--commit", action="store_true")
    a = ap.parse_args()

    decisions = load_decisions(a.xlsx)
    if not decisions:
        print("No confirm/reject decisions in the sheet (all blank/pending). Nothing to do.")
        return

    # index items by key
    idx = {}
    for f in ci.files():
        t = Path(f).read_text(encoding="utf-8")
        k = ci.key_of(t)
        if k:
            idx[k] = (f, t)

    plan, skips = [], []
    for key, dec in decisions:
        if dec.startswith("__invalid"):
            skips.append((key, f"invalid Decision value '{dec.split(':',1)[1]}' (use confirm/reject)"))
            continue
        if key not in idx:
            skips.append((key, "no item with this key"))
            continue
        f, t = idx[key]
        flag = ci.inferred_flag(t)
        if not flag:
            skips.append((key, "not inferred (already decided or never inferred) -- skipped"))
            continue
        if dec == "confirm":
            new, note = ci.apply_confirm(t, flag)
        else:
            new, note = ci.apply_reject(t, flag)
        plan.append((key, dec, f, t, new, note, flag))

    mode = "COMMIT" if a.commit else "DRY-RUN"
    print(f"{mode} | decisions in sheet: {len(decisions)} | to apply: {len(plan)} | skipped: {len(skips)}\n")
    for key, reason in skips:
        print(f"  skip {key}: {reason}")
    if skips:
        print()
    for key, dec, f, t, new, note, flag in plan:
        print(f"  {key:8} {dec:8} [{flag}] -> {note}")

    if not a.commit:
        v = ci.audit_verify()
        print(f"\nchain: ok={v.get('ok')} length={v.get('length')}")
        print("dry-run only. --commit to write.")
        return

    v = ci.audit_verify()
    if v.get("ok") is False:
        sys.exit("ABORT: audit chain broken -- refusing to write")
    n = 0
    for key, dec, f, t, new, note, flag in plan:
        shutil.copy2(f, f + ".pre-confirm")
        Path(f).write_text(new, encoding="utf-8")
        ci.audit_record({
            "action": "CONFIRM_ITEM" if dec == "confirm" else "REJECT_ITEM",
            "status": "OK", "redaction": "NONE", "dest": "portfolio",
            "file": Path(f).name, "key": key, "from_flag": flag,
            "note": note, "via": "round-trip",
        })
        n += 1
    print(f"\napplied {n} decision(s) with .pre-confirm backups + audit entries.")

if __name__ == "__main__":
    main()
