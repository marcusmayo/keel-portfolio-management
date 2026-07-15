#!/usr/bin/env python3
"""Multi-source reconcile worklist. Runs BOTH reconcile passes (backlog vs
Keel, Jira vs Keel) via subprocess, reads each reconcile.json before the next
overwrites it, and merges bucket rows by keel_key so one row shows what each
source asserts against a portfolio item.

Provenance is explicit: every source assertion is tagged BACKLOG or JIRA with
what it says (verdict). Portfolio items are tagged by origin (keel-origin vs
jira-linked, from source.ref). Source rows with no Keel match list separately.

Step 1 build: driver + merge only. Console summary; no xlsx yet. Read-only.
"""
import json, subprocess, sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Status equivalence: two trackers "disagree" only if they share no canonical class.
# Ruling: backlog / IP / in-progress / analysis / to-do / NYS are all ACTIVE (synonymous).
_CANON = {
    "done": "DONE", "completed": "DONE", "complete": "DONE", "released": "DONE",
    "dev verified": "DONE",
    "backlog": "ACTIVE", "ip": "ACTIVE", "in progress": "ACTIVE", "in-progress": "ACTIVE",
    "analysis": "ACTIVE", "in analysis": "ACTIVE", "to do": "ACTIVE", "nys": "ACTIVE",
    "not-started": "ACTIVE", "not started": "ACTIVE", "open": "ACTIVE",
    "requirement gathering": "ACTIVE", "ready": "ACTIVE", "in-review": "ACTIVE",
    "blocked": "BLOCKED",
}
def canon_status(tok):
    return _CANON.get((tok or "").strip().lower(), "OTHER:" + (tok or "").strip())

ROOT = Path(__file__).resolve().parent.parent
RECON = ROOT / "state" / "normalized" / "reconcile.json"
STATE = ROOT / "state"

def run_pass(source):
    """Run one reconcile pass, return its parsed reconcile.json."""
    r = subprocess.run(
        ["python3", str(ROOT / "tools" / "reconcile.py"), source],
        cwd=str(ROOT), capture_output=True, text=True, timeout=120)
    if r.returncode != 0:
        sys.exit(f"ERROR: reconcile [{source}] failed:\n{r.stderr[-500:]}")
    return json.loads(require(RECON).read_text(encoding="utf-8"))

import re, glob as _glob
from _require import require
_ORIGIN_INDEX = None
def _build_origin_index():
    """Index every portfolio + support YAML by its internal key: field ->
    origin. Filenames are name-slugs (epic-foo.yaml), so the key lives in the
    file body, not the filename. Source ref present -> jira-linked, else keel-origin."""
    idx = {}
    dirs = [STATE, ROOT / "support"]
    for d in dirs:
        for f in _glob.glob(str(d / "*.yaml")):
            if Path(f).stem.startswith("_"):
                continue
            txt = Path(f).read_text(encoding="utf-8")
            km = re.search(r"^\s*key:\s*['\"]?([A-Z]+-\d+)", txt, re.M)
            if not km:
                continue
            has_ref = bool(re.search(r"^\s*source:.*ref:\s*['\"]?[A-Z]+-\d+", txt, re.M))
            idx[km.group(1)] = "jira-linked" if has_ref else "keel-origin"
    return idx
def keel_origin(key):
    global _ORIGIN_INDEX
    if _ORIGIN_INDEX is None:
        _ORIGIN_INDEX = _build_origin_index()
    return _ORIGIN_INDEX.get(key, "unknown")

_REF_INDEX = None
def jira_key(key):
    """Source ref for a portfolio item, blank if keel-origin/unmatched."""
    global _REF_INDEX
    if _REF_INDEX is None:
        _REF_INDEX = {}
        for f in _glob.glob(str(STATE / "*.yaml")) + _glob.glob(str(ROOT / "support" / "*.yaml")):
            if Path(f).stem.startswith("_"):
                continue
            txt = Path(f).read_text(encoding="utf-8")
            km = re.search(r"^\s*key:\s*['\"]?([A-Z]+-\d+)", txt, re.M)
            rm = re.search(r"^\s*source:.*ref:\s*['\"]?([A-Z]+-\d+)", txt, re.M)
            if km:
                _REF_INDEX[km.group(1)] = rm.group(1) if rm else ""
    return _REF_INDEX.get(key, "")

_SCORES = None
def score3(key):
    """(wsjf, rice, grounding) display triple for a keel key; blanks when unscored."""
    global _SCORES
    if _SCORES is None:
        import glob as _sg
        _SCORES = {}
        for _f in _sg.glob(str(STATE / "*.yaml")) + _sg.glob(str(ROOT / "support" / "*.yaml")):
            if Path(_f).stem.startswith("_"):
                continue
            _t = Path(_f).read_text(encoding="utf-8")
            _km = re.search(r'^\s*key:\s*"?([A-Z]+-\d+)', _t, re.M)
            if not _km:
                continue
            _wm = re.search(r'^\s*wsjf:\s*\{.*?score:\s*([0-9.]+)\s*\}', _t, re.M)
            _rm = re.search(r'^\s*rice:\s*\{.*?score:\s*([0-9.]+)\s*\}', _t, re.M)
            _gm = re.search(r'^\s*scored:\s*\{.*?grounding:\s*([a-z]+)', _t, re.M)
            _SCORES[_km.group(1)] = (float(_wm.group(1)) if _wm else "",
                                     float(_rm.group(1)) if _rm else "",
                                     _gm.group(1) if _gm else "")
    return _SCORES.get(key, ("", "", ""))

def main():
    backlog = run_pass("backlog")
    jira = run_pass("jira")
    print(f"backlog buckets: { {k: len(v) for k,v in backlog['buckets'].items()} }")
    print(f"jira buckets:    { {k: len(v) for k,v in jira['buckets'].items()} }")

    # Merge by keel_key. Each portfolio item accumulates assertions from both sources.
    merged = {}   # keel_key -> {keel_name, keel_status, assertions: [(SOURCE, verdict, src_name, src_ref)]}
    source_only = []  # rows with no keel_key (source-declared, no portfolio match)

    for src_tag, data in (("BACKLOG", backlog), ("JIRA", jira)):
        for bucket, rows in data["buckets"].items():
            for row in rows:
                kk = (row.get("keel_key") or "").strip()
                assertion = (src_tag, row.get("verdict", bucket),
                             row.get("src_name", ""), row.get("src_ref", ""),
                             row.get("src_status", ""))
                if kk:
                    if kk not in merged:
                        merged[kk] = {"keel_name": row.get("keel_name", ""),
                                      "keel_status": row.get("keel_status", ""),
                                      "assertions": []}
                    if assertion not in merged[kk]["assertions"]:
                        merged[kk]["assertions"].append(assertion)
                else:
                    source_only.append((src_tag, bucket, row.get("verdict", ""),
                                        row.get("src_name", ""), row.get("src_ref", ""),
                                        row.get("src_status", "")))

    both = {k: v for k, v in merged.items() if len(set(a[0] for a in v["assertions"])) > 1}
    print(f"\nmerged portfolio items with assertions: {len(merged)}")
    print(f"  touched by BOTH sources: {len(both)}")
    print(f"source-only rows (no Keel match): {len(source_only)}")

    HFILL = PatternFill("solid", fgColor="1F3B1B")
    HFONT = Font(bold=True, color="FFFFFF")
    WARN = PatternFill("solid", fgColor="C0392B")
    def header(ws, cols):
        for c, label in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=label)
            cell.fill = HFILL; cell.font = HFONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def asserts_str(assertions, tag):
        parts = [f"{a[1]}({a[4] or '-'})" for a in assertions if a[0] == tag]
        return "; ".join(parts)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Cross-Source (both + single, all merged items), disagreement flagged
    ws = wb.create_sheet("Cross-Source")
    cols = ["Keel Key", "Jira Key", "Keel Item", "Keel Status", "Origin",
            "Backlog Asserts", "Jira Asserts", "Sources", "Disagree?", "WSJF", "RICE", "Grounding"]
    header(ws, cols)
    def status_classes(assertions, tag):
        return set(canon_status(a[4]) for a in assertions if a[0] == tag and a[4])
    r = 1
    ordered_items = list(both.items()) + [(k, v) for k, v in merged.items() if k not in both]
    for kk, v in ordered_items:
        r += 1
        b_str = asserts_str(v["assertions"], "BACKLOG")
        j_str = asserts_str(v["assertions"], "JIRA")
        srcs = "+".join(sorted(set(a[0] for a in v["assertions"])))
        bc = status_classes(v["assertions"], "BACKLOG")
        jc = status_classes(v["assertions"], "JIRA")
        # disagree only when both sources assert AND share no canonical class
        disagree = "YES" if (bc and jc and not (bc & jc)) else ""
        vals = [kk, jira_key(kk), v["keel_name"], v["keel_status"], keel_origin(kk),
                b_str, j_str, srcs, disagree] + list(score3(kk))
        for c, val in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=val)
        if disagree:
            ws.cell(row=r, column=9).fill = WARN
            ws.cell(row=r, column=9).font = Font(bold=True, color="FFFFFF")
    for c, w in zip(range(1, 13), [10, 10, 36, 13, 12, 32, 32, 12, 10, 8, 8, 10]):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(r,1)+1}"

    # Sheet 2: Source-Only (no Keel match)
    ws2 = wb.create_sheet("Source-Only")
    cols2 = ["Source", "Bucket", "Verdict", "Source Item", "Source Ref", "Source Status"]
    header(ws2, cols2)
    for i, row in enumerate(source_only, 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=i, column=c, value=val)
    for c, w in zip(range(1, 7), [10, 12, 14, 42, 12, 14]):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:F{len(source_only)+1}"

    # Sheet 3: Keel-Origin = ALL blank-ref items (the To-Author population),
    # pulled from the origin index directly so items no source touches still appear.
    ws3 = wb.create_sheet("Keel-Origin")
    header(ws3, ["Keel Key", "Type", "Keel Item", "Status", "Parent Key", "Parent Name", "Description (draft)", "Acceptance Criteria (draft)", "Asserted by", "WSJF", "RICE", "Grounding"])
    if _ORIGIN_INDEX is None:
        keel_origin("")  # force index build
    def gd_desc(txt):
        """Block-scalar-aware description reader: single-line quoted OR
        `description: >`/`|` continuation lines joined."""
        dl = txt.splitlines()
        for i, l in enumerate(dl):
            m = re.match(r'^  description:\s*"(.*)"\s*$', l)
            if m:
                return m.group(1)
            if re.match(r'^  description:\s*[>|]', l):
                buf = []
                for l2 in dl[i + 1:]:
                    if l2.strip() == "":
                        continue
                    if re.match(r'^    \S|^     ', l2):
                        buf.append(l2.strip())
                    else:
                        break
                return " ".join(buf)
            if re.match(r'^  description:\s*$', l):
                return ""
        return ""

    def _rejected(_key):
        import glob as _rg
        for _rf in _rg.glob(str(STATE / "*.yaml")) + _rg.glob(str(ROOT / "support" / "*.yaml")):
            _rt = Path(_rf).read_text(encoding="utf-8")
            if re.search(rf'^\s*key:\s*"?{re.escape(_key)}"?', _rt, re.M):
                return bool(re.search(r'^    status:\s*"?rejected"?', _rt, re.M))
        return False
    ko_keys = sorted(k for k, o in _ORIGIN_INDEX.items() if o == "keel-origin" and not _rejected(k))
    # pull name/status/type/parent from each item's YAML by key
    import glob as _g
    meta = {}
    for f in _g.glob(str(STATE / "*.yaml")) + _g.glob(str(ROOT / "support" / "*.yaml")):
        if Path(f).stem.startswith("_"):
            continue
        txt = Path(f).read_text(encoding="utf-8")
        km = re.search(r"^\s*key:\s*['\"]?([A-Z]+-\d+)", txt, re.M)
        if not km:
            continue
        g = lambda field: (re.search(rf"^\s*{field}:\s*['\"]?([^'\"\n]+)", txt, re.M) or [None, ""])[1].strip()
        # acceptance_criteria is a YAML list: capture bullet lines until the next top-level key
        ac_items = []
        acm = re.search(r"^\s*acceptance_criteria:\s*(.*?)(?=^\s*\w[\w-]*:|\Z)", txt, re.M | re.S)
        if acm:
            for line in acm.group(1).splitlines():
                ls = line.strip()
                if ls.startswith("- "):
                    ac_items.append(ls[2:].strip().strip("'\""))
        meta[km.group(1)] = {"name": g("name"), "status": g("status"),
                             "type": g("type"), "parent": g("parent"),
                             "description": (re.sub(r"^\s*\[\s*draft\s*[\u2014\u2013-]\s*review\s*\]\s*", "", gd_desc(txt)) if re.search(r"^\s*source:.*ref:\s*[A-Z]+-\d+", txt, re.M) else gd_desc(txt)),
                             "ac": "\n".join(ac_items)}
    for i, kk in enumerate(ko_keys, 2):
        m = meta.get(kk, {})
        srcs = "+".join(sorted(set(a[0] for a in merged[kk]["assertions"]))) if kk in merged else "none"
        pkey = m.get("parent", "")
        pmeta = meta.get(pkey, {}) if pkey else {}
        vals = [kk, m.get("type", ""), m.get("name", ""), m.get("status", ""),
                pkey, pmeta.get("name", ""),
                m.get("description", ""), m.get("ac", ""), srcs] + list(score3(kk))
        for c, val in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=c, value=val)
            if c in (7, 8):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for c, w in zip(range(1, 13), [10, 8, 34, 12, 10, 26, 50, 50, 12, 8, 8, 10]):
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:L{len(ko_keys)+1}"

    OUTDIR = ROOT / "exports"
    OUTDIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now()
    out = OUTDIR / f"multisource-{ts.strftime('%Y-%m-%d')}_{ts.strftime('%H%M')}.xlsx"
    # Sheet: Unconfirmed -- LLM-inferred items (flag: draft-inferred*) awaiting
    # operator decision. Column names deliberately DO NOT match the round-trip
    # edit apply's want-sets (key/description), so apply_portfolio_edits can
    # never select this sheet. Decision column: operator writes confirm|reject.
    wsU = wb.create_sheet("Unconfirmed")
    ucols = ["Unconfirmed Key", "Type", "Item Name", "Variant", "Current Status",
             "Inference Description", "Decision (confirm/reject)"]
    header(wsU, ucols)
    import glob as _gu
    _dp_u = re.compile(r"\[\s*draft\s*[\u2014\u2013-]\s*review\s*\]\s*", re.I)
    urow = 1
    for _f in sorted(_gu.glob(str(STATE / "*.yaml")) + _gu.glob(str(ROOT / "support" / "*.yaml"))):
        if Path(_f).stem.startswith("_"):
            continue
        _t = Path(_f).read_text(encoding="utf-8")
        _fm = re.search(r'^\s*flag:\s*"?(draft-inferred(?:-[a-z-]+)?)"?\s*$', _t, re.M)
        if not _fm:
            continue
        _flag = _fm.group(1)
        _variant = _flag[len("draft-inferred"):].lstrip("-") or "(plain)"
        _k = (re.search(r'^\s*key:\s*"?([A-Z]+-\d+)', _t, re.M) or [None, "?"])[1]
        _typ = (re.search(r'^  type:\s*(\w+)', _t, re.M) or [None, "?"])[1]
        _name = (re.search(r'^  name:\s*"?(.*?)"?\s*$', _t, re.M) or [None, ""])[1]
        _st = (re.search(r'^    status:\s*"?(\w+)', _t, re.M) or [None, "?"])[1]
        _desc = _dp_u.sub("", gd_desc(_t)).strip()
        urow += 1
        for _c, _v in enumerate([_k, _typ, _name, _variant, _st, _desc, ""], 1):
            _cell = wsU.cell(row=urow, column=_c, value=_v)
            _cell.alignment = Alignment(vertical="top", wrap_text=True)
    for _c, _w in zip(range(1, 8), [16, 8, 30, 18, 14, 60, 24]):
        wsU.column_dimensions[get_column_letter(_c)].width = _w
    wsU.freeze_panes = "A2"
    wsU.auto_filter.ref = f"A1:G{urow}"

    # Sheet: Semantic Matches -- LLM SAME/DISTINCT verdicts on Jira<->Keel pairs
    # (from reconcile_semantic.py, read from disk). Backlog is NOT judged here;
    # backlog reconciliation lives in Cross-Source (Disagree?). Column names
    # deliberately avoid Key/Description so round-trip apply tools never select it.
    # Visibility only (Option A): to act, use /merge-accept, /merge-reject, /merge-distinct.
    wsS = wb.create_sheet("Semantic Matches")
    scols = ["Jira Ref", "Jira Name", "Keel Match", "Keel Item Name",
             "Deterministic Verdict", "Semantic Verdict", "Semantic Reason"]
    header(wsS, scols)
    srow = 1
    # Read durable semantic verdicts (written by reconcile_semantic.py). Reading a
    # SEPARATE file, not reconcile.json, because this export re-runs reconcile
    # (run_pass) which rewrites reconcile.json and would wipe the verdicts.
    try:
        _sdata = json.loads(Path("state/normalized/semantic.json").read_text(encoding="utf-8"))
        _spairs = _sdata.get("pairs", [])
    except Exception:
        _spairs = []
    for _e in _spairs:
        srow += 1
        _vals = [_e.get("src_ref", ""), _e.get("src_name", ""),
                 _e.get("keel_key", ""), _e.get("keel_name", ""),
                 _e.get("verdict", ""), _e.get("semantic_verdict", ""), _e.get("semantic_reason", "")]
        for _c, _v in enumerate(_vals, 1):
            _cell = wsS.cell(row=srow, column=_c, value=_v)
            _cell.alignment = Alignment(vertical="top", wrap_text=True)
    for _c, _w in zip(range(1, 8), [10, 34, 12, 30, 16, 16, 60]):
        wsS.column_dimensions[get_column_letter(_c)].width = _w
    wsS.freeze_panes = "A2"
    wsS.auto_filter.ref = f"A1:G{srow}"
    # Decision column (G) confirm/reject dropdown -- minimal validation.
    _dv = DataValidation(type="list", formula1='"confirm,reject"', allow_blank=True)
    wsU.add_data_validation(_dv)
    _dv.add(f"G2:G{urow}")

    # Legend sheet -- appended last; has no Key/Description columns so the
    # round-trip apply tool (find_sheet_and_cols) can never select it.
    wsL = wb.create_sheet("Legend")
    for _i, _row in enumerate([
        ("Column", "How to read"),
        ("", ""),
        ("--- SCORES ---", ""),
        ("WSJF", "Cost of delay / job size. Higher = do sooner."),
        ("RICE", "Reach x Impact x Confidence / Effort (person-weeks). Higher = better."),
        ("Grounding", "grounded = traced to item/context text; generic = pattern estimate (review harder)."),
        ("Blank scores", "Item unscored, needs-input, or a support bug (no prioritization block)."),
        ("", ""),
        ("--- SHEETS ---", ""),
        ("Cross-Source", "Items with a source (backlog and/or Jira). Disagree? column flags status conflicts. Read-only reference."),
        ("Source-Only", "Items in a source with NO Keel item yet. To bring in: run the import lane (/reconcile-run -> /export -> /apply -> /apply-commit)."),
        ("Keel-Origin", "Keel-authored items (no Jira ref). This is where you review/approve DRAFT descriptions + AC. See apply-edits below."),
        ("Unconfirmed", "LLM-INFERRED items Keel guessed from notes. NOT yet real work until you decide. See apply-inference below."),
        ("", ""),
        (">> APPLY-EDITS (content)", "Question: is the drafted TEXT good enough to be the real content?"),
        ("  Which sheet", "Keel-Origin (or the Portfolio export). Affects description + acceptance criteria text."),
        ("  How to approve", "DELETE the [draft - review] prefix from a Description or an AC line you approve. Leave the prefix to keep it pending."),
        ("  Re-upload -> ", "Reply /apply-edits. Writes the de-prefixed text into the item YAML."),
        ("  Note", "Jira-origin items are NOT editable here (Jira owns their content) -- they are skipped."),
        ("", ""),
        (">> APPLY-INFERENCE (legitimacy)", "Question: should this inferred item EXIST as tracked work at all?"),
        ("  Which sheet", "Unconfirmed. Affects the draft-inferred flag (whether the item is confirmed real or rejected)."),
        ("  How to decide", "In the Decision column type exactly confirm or reject (lowercase). Leave blank to keep it pending."),
        ("  Re-upload -> ", "Reply /apply-inference. Applies your confirm/reject decisions."),
        ("  On confirm", "draft-inferred flag cleared -> item becomes confirmed portfolio work. Residual needs-repro / needs-decomposition is kept."),
        ("  On reject", "status set to rejected; item filtered OUT of Cross-Source / Keel-Origin working views (kept for the record, reversible)."),
        ("  Variant meaning", "(plain) = inferred item; needs-repro = inferred bug not yet reproduced; needs-decomposition = inferred but too big."),
        ("", ""),
        ("", ""),
        (">> SEMANTIC MATCHES (Jira<->Keel)", "LLM SAME/DISTINCT judgments on Jira<->Keel pairs. Proposals only -- nothing is auto-merged."),
        ("  Scope", "Jira vs Keel (ST/EP) pairs ONLY. Backlog is NOT judged here -- backlog conflicts are in Cross-Source (Disagree?)."),
        ("  SAME means", "The LLM thinks these are the same capability/work. A merge candidate."),
        ("  DISTINCT means", "The LLM thinks these are different work despite a title/ref match. Keep separate."),
        ("  How to act", "Agree with SAME -> run /merge-accept. Reject a pair -> /merge-reject <keys> or /merge-distinct <keys>."),
        ("  Note", "This sheet does not merge anything itself. It surfaces verdicts so you can decide, then act via the merge commands."),
        ("", ""),
        ("--- ORDER OF WORK ---", "1) apply-inference to confirm an item is real. 2) apply-edits to approve its draft text. Legitimacy first, then content."),
    ], 1):
        for _c, _v in enumerate(_row, 1):
            wsL.cell(row=_i, column=_c, value=_v)
    wsL.column_dimensions["A"].width = 14
    wsL.column_dimensions["B"].width = 95
    wb.save(out)
    disagreements = sum(1 for kk, v in both.items()
                        if set(a[1] for a in v["assertions"] if a[0]=="BACKLOG")
                        != set(a[1] for a in v["assertions"] if a[0]=="JIRA"))
    print(f"\nwrote {out}  (cross-source={len(merged)}  both={len(both)}  disagree={disagreements}  source-only={len(source_only)}  keel-origin={len([k for k,o in (_ORIGIN_INDEX or {}).items() if o=='keel-origin'])})")

if __name__ == "__main__":
    main()
