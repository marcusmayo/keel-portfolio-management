#!/usr/bin/env python3
"""Generate the ADO mirror of the Northwind BACKLOG corpus for the parity test.

Reads the generated knowledge/import/raw/Northwind Backlog.xlsx and transforms each
row into an ADO CSV row that normalizes to the SAME canonical (type, status) as the
backlog lane -- so reconcile parity between the ADO lane and the backlog lane can be
asserted on equivalent content (not on unrelated data).

Deliberate fidelity choice: the ADO mirror carries NO cross-tracker (NWR) key. A real
ADO CSV export has no notes-embedded-key field, so the backlog's ref-only plants
(title-drift NWR-122, embedded-key NWR-142) land by TITLE in the ADO lane. That
difference is exactly what the parity test measures and documents -- it is a property
of the source format, not a defect.

Regenerate:  python3 tests/fixtures/ado/gen_northwind_ado.py
             (requires the corpus: python3 examples/northwind/gen_corpus.py .)
"""
import csv
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
XLSX = ROOT / "knowledge" / "import" / "raw" / "Northwind Backlog.xlsx"
OUT = Path(__file__).resolve().parent / "ado_northwind_backlog.csv"

# backlog Type -> ADO Work Item Type
TYPE = {"epic": "Epic", "feature": "Feature", "story": "User Story"}
# backlog Status code -> ADO State, chosen so each maps to the SAME canonical value
# under the custom state vocab the parity test supplies via ADO_STATE_MAP:
#   New->not-started  Active->in-progress  Closed->done  Blocked->blocked  Analysis->analysis
STATE = {"IP": "Active", "NYS": "New", "DONE": "Closed", "BLOCKED": "Blocked",
         "IN ANALYSIS": "Analysis", "NEEDS ANALYSIS": "Analysis", "DUPLICATE": "Removed"}


def main():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr)}
    out = [["ID", "Work Item Type", "Title", "Assigned To", "State", "Tags"]]
    wid = 50001
    for r in rows[1:]:
        feature = str(r[col["Feature"]] or "").strip()
        if not feature:
            continue
        btype = str(r[col["Type"]] or "").strip().lower()
        bstat = str(r[col["Status"]] or "").strip().upper()
        owner = str(r[col["Owner"]] or "").strip()
        out.append([str(wid), TYPE.get(btype, btype or "Issue"), feature,
                    owner, STATE.get(bstat, "New"), ""])
        wid += 1
    with OUT.open("w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(out)
    print(f"wrote {OUT.name}: {len(out) - 1} ADO rows mirroring the Northwind backlog")


if __name__ == "__main__":
    main()
